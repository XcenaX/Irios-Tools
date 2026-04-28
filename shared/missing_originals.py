from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import re
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, PatternFill, Side

from tools.create_report_template import HEADER_IMAGE, build_template, sheet_range_size_pixels
from .missing_originals_contract import (
    PRODUCT_CODE,
    PRODUCT_VERSION,
    build_report_data_from_payload,
)


TOKEN_FILE_NAME = ".missing_originals_license.json"
DEFAULT_API_BASE_URL = "http://127.0.0.1:8000"
DEFAULT_PRODUCT_NAME = "Missing Originals"


@dataclass(frozen=True)
class SourceConfig:
    kind: str
    original_names: tuple[str, ...]
    date_col: str
    counterparty_col: str
    amount_col: str
    account_number_names: tuple[str, ...]
    operation_names: tuple[str, ...] = ()
    operation_index: int | None = None
    account_number_index: int | None = None
    comment_names: tuple[str, ...] = ()
    comment_index: int | None = None
    avr_names: tuple[str, ...] = ()
    avr_index: int | None = None


@dataclass(frozen=True)
class BuildResult:
    receipts_count: int
    sales_count: int
    output_file: Path
    company_name: str

    @property
    def total_count(self) -> int:
        return self.receipts_count + self.sales_count


@dataclass(frozen=True)
class MultiBuildResult:
    company_name: str
    builds: tuple[BuildResult, ...]

    @property
    def total_files(self) -> int:
        return len(self.builds)

    @property
    def total_receipts_count(self) -> int:
        return sum(build.receipts_count for build in self.builds)

    @property
    def total_sales_count(self) -> int:
        return sum(build.sales_count for build in self.builds)

    @property
    def total_count(self) -> int:
        return sum(build.total_count for build in self.builds)


RECEIPTS_CONFIG = SourceConfig(
    kind="Поступление работ и услуг",
    original_names=("Оригинал документа (Поступления ТМЗ и услуг)",),
    date_col="Дата",
    counterparty_col="Контрагент",
    amount_col="Сумма",
    operation_names=("Вид операции",),
    operation_index=11,
    account_number_names=("Номер",),
    account_number_index=6,
    comment_names=("Комментарий",),
    comment_index=20,
    avr_names=("АВР",),
    avr_index=17,
)

SALES_CONFIG = SourceConfig(
    kind="Реализация работ и услуг",
    original_names=("Оригинал", "Оригинал документов (Реализации ТМЗ и услуг)"),
    date_col="Дата",
    counterparty_col="Контрагент",
    amount_col="Сумма",
    operation_names=("Вид операции",),
    operation_index=6,
    account_number_names=("Номер",),
    account_number_index=6,
    comment_names=("Комментарий",),
    comment_index=18,
    avr_names=("ЭАВР", "АВР"),
    avr_index=15,
)


THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LAST_TABLE_COLUMN = 10
RECEIPTS_TITLE_ROW = 4
RECEIPTS_FIRST_DATA_ROW = 6
RECEIPTS_TOTAL_ROW = 12
TEMPLATE_DATA_ROWS = 6


def collapse_spaces(value: str) -> str:
    return " ".join(value.split())


def normalize_company_spacing(value: str) -> str:
    value = collapse_spaces(value)
    value = re.sub(r"\s*-\s*", "-", value)
    return value.strip()


def sanitize_path_component(value: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|]+', " ", value)
    cleaned = normalize_company_spacing(cleaned).strip(" .-_")
    return cleaned or "Компания"


def extract_comment_group(value: str) -> str | None:
    match = re.fullmatch(r"\s*(\d+)\s*", value)
    if not match:
        return None
    return match.group(1)


def extract_company_hint(path: Path) -> str:
    value = path.stem
    patterns = [
        r"(?i)\bнет\s+оригиналов?\b",
        r"(?i)\bпоступлени[еяй]\b",
        r"(?i)\bреализаци[яй]\b",
        r"(?i)\bработ\b",
        r"(?i)\bи\b",
        r"(?i)\bуслуг\b",
        r"(?i)\bза\b",
        r"(?i)\bгод\b",
        r"(?i)\bотчет\b",
        r"(?i)\bотч[её]т\b",
        r"(?i)\b\d{4}\b",
        r"_+",
    ]
    for pattern in patterns:
        value = re.sub(pattern, " ", value)
    return normalize_company_spacing(value)


def common_words_name(first: str, second: str) -> str:
    second_words = {word.casefold() for word in second.split()}
    words = [word for word in first.split() if word.casefold() in second_words]
    return normalize_company_spacing(" ".join(words))


def normalize_yes_no(value: object) -> str:
    return str(value or "").strip().casefold()


def normalize_text(value: object) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_header_text(value: object) -> str:
    text = str(value or "").strip().casefold().replace("ё", "е")
    text = re.sub(r"\s+", " ", text)
    return text


def resolve_column_name(
    df: pd.DataFrame,
    *,
    preferred_names: tuple[str, ...],
    fallback_index: int | None,
    label: str,
    path: Path,
    required: bool,
    contains_keywords: tuple[str, ...] = (),
) -> str | None:
    for name in preferred_names:
        if name in df.columns:
            return name

    normalized_keywords = tuple(normalize_header_text(keyword) for keyword in contains_keywords if keyword)
    if normalized_keywords:
        for column in df.columns:
            normalized_column = normalize_header_text(column)
            if all(keyword in normalized_column for keyword in normalized_keywords):
                return str(column)

    if fallback_index is not None and len(df.columns) >= fallback_index:
        return str(df.columns[fallback_index - 1])

    if required:
        expected_names = ", ".join(preferred_names) if preferred_names else "-"
        raise ValueError(
            f"В файле '{path.name}' не найдена колонка '{label}'. "
            f"Искал названия: {expected_names}; резервный номер колонки: {fallback_index}."
        )
    return None


def get_column_series(
    df: pd.DataFrame,
    *,
    preferred_names: tuple[str, ...],
    fallback_index: int | None,
    label: str,
    path: Path,
    required: bool,
    contains_keywords: tuple[str, ...] = (),
) -> pd.Series:
    column_name = resolve_column_name(
        df,
        preferred_names=preferred_names,
        fallback_index=fallback_index,
        label=label,
        path=path,
        required=required,
        contains_keywords=contains_keywords,
    )
    if column_name is None:
        return pd.Series([""] * len(df), index=df.index, dtype="object")
    return df[column_name]


def read_source(path: Path, config: SourceConfig) -> pd.DataFrame:
    df = pd.read_excel(path)

    original_col = resolve_column_name(
        df,
        preferred_names=config.original_names,
        fallback_index=None,
        label="Оригинал",
        path=path,
        required=True,
        contains_keywords=("оригинал",),
    )
    required = [config.date_col, config.counterparty_col, config.amount_col]
    missing = [col for col in required if col not in df.columns]
    if missing or original_col is None:
        missing_labels = [*missing]
        if original_col is None:
            missing_labels.append("Оригинал")
        raise ValueError(
            f"В файле '{path.name}' не найдены колонки: {', '.join(missing_labels)}"
        )

    report_df = pd.DataFrame(
        {
            "original": df[original_col],
            "date": df[config.date_col],
            "counterparty": df[config.counterparty_col],
            "amount": df[config.amount_col],
            "operation_type": get_column_series(
                df,
                preferred_names=config.operation_names,
                fallback_index=config.operation_index,
                label="Вид операции",
                path=path,
                required=False,
            ),
            "account_number": get_column_series(
                df,
                preferred_names=config.account_number_names,
                fallback_index=config.account_number_index,
                label="Номер",
                path=path,
                required=True,
            ),
            "avr": get_column_series(
                df,
                preferred_names=config.avr_names,
                fallback_index=config.avr_index,
                label="АВР/ЭАВР",
                path=path,
                required=False,
            ),
            "comment": get_column_series(
                df,
                preferred_names=config.comment_names,
                fallback_index=config.comment_index,
                label="Комментарий",
                path=path,
                required=True,
            ),
        }
    ).copy()

    report_df["original"] = report_df["original"].map(normalize_yes_no)
    report_df = report_df[report_df["original"] == "нет"].copy()

    report_df["avr"] = report_df["avr"].map(normalize_text)
    report_df = report_df[report_df["avr"] == ""].copy()

    report_df["comment"] = report_df["comment"].map(normalize_text).str.casefold()
    report_df = report_df[report_df["comment"] != "гпх"].copy()

    report_df["comment_group"] = report_df["comment"].map(extract_comment_group)
    report_df["date"] = pd.to_datetime(report_df["date"], errors="coerce", dayfirst=True)
    report_df["counterparty"] = report_df["counterparty"].fillna("").astype(str).str.strip()
    report_df["amount"] = pd.to_numeric(report_df["amount"], errors="coerce")
    report_df["operation_type"] = report_df["operation_type"].map(normalize_text)
    report_df["account_number"] = report_df["account_number"].map(normalize_text)
    report_df = report_df.dropna(subset=["date", "counterparty", "amount"])
    report_df = report_df.sort_values(
        ["date", "counterparty", "amount", "operation_type", "account_number"]
    ).reset_index(drop=True)

    report_df = report_df[
        ["date", "counterparty", "amount", "operation_type", "account_number", "comment_group"]
    ]
    report_df.insert(0, "index", range(1, len(report_df) + 1))
    report_df.insert(0, "section", config.kind)
    return report_df


def detect_company_name(receipts_file: Path, sales_file: Path) -> str:
    receipts_hint = extract_company_hint(receipts_file)
    sales_hint = extract_company_hint(sales_file)

    if receipts_hint and sales_hint:
        if receipts_hint.casefold() == sales_hint.casefold():
            return receipts_hint
        common_name = common_words_name(receipts_hint, sales_hint)
        if common_name:
            return common_name

    if receipts_hint:
        return receipts_hint
    if sales_hint:
        return sales_hint
    return "Компания"


def detect_company_from_sources(receipts_file: Path, sales_file: Path) -> str | None:
    for path in (receipts_file, sales_file):
        try:
            df = pd.read_excel(path)
        except Exception:
            continue

        org_col = resolve_column_name(
            df,
            preferred_names=("Организация",),
            fallback_index=None,
            label="Организация",
            path=path,
            required=False,
        )
        if not org_col:
            continue

        series = df[org_col].dropna().astype(str).map(normalize_company_spacing)
        series = series[series.astype(str).str.strip() != ""]
        if not series.empty:
            return series.iloc[0]
    return None


def resolve_report_company_name(receipts_file: Path, sales_file: Path) -> str:
    return detect_company_from_sources(receipts_file, sales_file) or detect_company_name(
        receipts_file, sales_file
    )


def build_company_output_path(base_dir: Path, company_name: str) -> Path:
    safe_company_name = sanitize_path_component(company_name)
    company_dir = base_dir / safe_company_name
    timestamp = datetime.now().strftime("%d.%m.%Y %H.%M")
    report_name = f"{safe_company_name} {timestamp}.xlsx"
    return company_dir / report_name


def make_unique_output_path(path: Path) -> Path:
    if not path.exists():
        return path

    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    counter = 1
    while True:
        candidate = parent / f"{stem} ({counter}){suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def format_period_label(*frames: pd.DataFrame) -> str:
    dates: list[datetime] = []
    for frame in frames:
        if frame.empty or "date" not in frame.columns:
            continue
        dates.extend(frame["date"].dropna().tolist())

    if not dates:
        return datetime.now().strftime("%d.%m.%Y")
    start_date = min(dates)
    end_date = max(dates)
    if start_date.date() == end_date.date():
        return start_date.strftime("%d.%m.%Y")
    return f"{start_date:%d.%m.%Y} - {end_date:%d.%m.%Y}"


def dataframe_to_rows(df: pd.DataFrame) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in df.to_dict(orient="records"):
        payload_row = {
            "index": int(row["index"]),
            "counterparty": str(row["counterparty"]),
            "date": pd.Timestamp(row["date"]).to_pydatetime().date().isoformat(),
            "amount": float(row["amount"]),
            "operation_type": str(row["operation_type"]),
            "account_number": str(row["account_number"]),
        }
        comment_group = row.get("comment_group")
        if pd.notna(comment_group) and comment_group not in ("", None):
            payload_row["comment_group"] = str(comment_group)
        rows.append(payload_row)
    return rows


def with_reindexed_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized_rows: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=1):
        cloned = dict(row)
        cloned["index"] = index
        normalized_rows.append(cloned)
    return normalized_rows


def split_report_rows_by_comment_group(
    report_data: dict[str, Any],
) -> list[tuple[str | None, dict[str, Any]]]:
    grouped_receipts: dict[str | None, list[dict[str, Any]]] = {}
    grouped_sales: dict[str | None, list[dict[str, Any]]] = {}

    for row in report_data["receipts_rows"]:
        grouped_receipts.setdefault(row.get("comment_group"), []).append(dict(row))
    for row in report_data["sales_rows"]:
        grouped_sales.setdefault(row.get("comment_group"), []).append(dict(row))

    group_keys = set(grouped_receipts) | set(grouped_sales)
    ordered_group_keys = [None]
    ordered_group_keys.extend(sorted(key for key in group_keys if key is not None))

    grouped_reports: list[tuple[str | None, dict[str, Any]]] = []
    for group_key in ordered_group_keys:
        receipts_rows = with_reindexed_rows(grouped_receipts.get(group_key, []))
        sales_rows = with_reindexed_rows(grouped_sales.get(group_key, []))
        if group_key is not None and not receipts_rows and not sales_rows:
            continue
        grouped_reports.append(
            (
                group_key,
                {
                    "company_name": report_data["company_name"],
                    "period_label": report_data["period_label"],
                    "receipts_rows": receipts_rows,
                    "sales_rows": sales_rows,
                },
            )
        )

    return grouped_reports


def build_group_output_path(base_output_file: Path, comment_group: str | None) -> Path:
    if comment_group is None:
        return base_output_file
    return base_output_file.with_name(
        f"{base_output_file.stem} - {sanitize_path_component(comment_group)}{base_output_file.suffix}"
    )


def build_payload_from_files(receipts_file: Path, sales_file: Path) -> dict[str, Any]:
    receipts = read_source(receipts_file, RECEIPTS_CONFIG)
    sales = read_source(sales_file, SALES_CONFIG)
    company_name = resolve_report_company_name(receipts_file, sales_file)
    period_label = format_period_label(receipts, sales)
    return {
        "company_name": company_name,
        "period_label": period_label,
        "receipts": dataframe_to_rows(receipts),
        "sales": dataframe_to_rows(sales),
    }


def clone_row_style(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for column in range(1, LAST_TABLE_COLUMN + 1):
        source = ws.cell(row=source_row, column=column)
        target = ws.cell(row=target_row, column=column)
        if source.has_style:
            target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def clear_section_rows(ws, start_row: int, row_count: int) -> None:
    for row_idx in range(start_row, start_row + row_count):
        for column in range(1, LAST_TABLE_COLUMN + 1):
            try:
                ws.cell(row=row_idx, column=column, value=None)
            except AttributeError:
                continue


def safe_unmerge(ws, cell_range: str) -> None:
    merged_ranges = {str(cell_range_ref) for cell_range_ref in ws.merged_cells.ranges}
    if cell_range in merged_ranges:
        ws.unmerge_cells(cell_range)


def drop_second_section(ws) -> None:
    sales_title_row = 14
    if ws.max_row >= sales_title_row:
        ws.delete_rows(sales_title_row, ws.max_row - sales_title_row + 1)


def merge_layout_blocks(ws, start_row: int, end_row: int) -> None:
    for row_idx in range(start_row, end_row + 1):
        safe_unmerge(ws, f"B{row_idx}:D{row_idx}")
        safe_unmerge(ws, f"G{row_idx}:I{row_idx}")
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=4)
        ws.merge_cells(start_row=row_idx, start_column=7, end_row=row_idx, end_column=9)


def attach_header_image(ws) -> None:
    if not HEADER_IMAGE.exists():
        return
    image = XLImage(str(HEADER_IMAGE))
    image.width, image.height = sheet_range_size_pixels(
        ws,
        start_col=1,
        end_col=10,
        start_row=1,
        end_row=3,
    )
    ws.add_image(image, "A1")


def write_section_to_template(
    ws,
    *,
    first_data_row: int,
    total_row: int,
    rows: list[dict[str, Any]],
    total_label: str,
) -> None:
    extra_rows = max(len(rows) - TEMPLATE_DATA_ROWS, 0)
    if extra_rows:
        ws.insert_rows(total_row, amount=extra_rows)
        for offset in range(extra_rows):
            clone_row_style(ws, total_row - 1, total_row + offset - 1)

    total_row += extra_rows
    row_count = max(TEMPLATE_DATA_ROWS, len(rows))
    clear_section_rows(ws, first_data_row, row_count)
    merge_layout_blocks(ws, first_data_row, first_data_row + row_count - 1)

    for row_offset, row in enumerate(rows):
        current_row = first_data_row + row_offset
        ws.cell(row=current_row, column=1, value=int(row["index"]))
        ws.cell(row=current_row, column=2, value=row["counterparty"])
        date_cell = ws.cell(
            row=current_row,
            column=5,
            value=datetime.fromisoformat(row["date"]),
        )
        amount_cell = ws.cell(row=current_row, column=6, value=float(row["amount"]))
        ws.cell(row=current_row, column=7, value=row["operation_type"])
        ws.cell(row=current_row, column=10, value=row["account_number"])
        date_cell.number_format = "dd.mm.yyyy"
        amount_cell.number_format = '#,##0 "₸"'

    total_amount = sum(float(row["amount"]) for row in rows)
    total_cell = ws.cell(
        row=total_row,
        column=1,
        value=f"ИТОГО: {total_label}: {len(rows)} документов / {total_amount:,.0f} ₸".replace(",", " "),
    )
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=LAST_TABLE_COLUMN)
    total_cell.alignment = Alignment(horizontal="left", vertical="center")


def render_report_data_to_excel(report_data: dict[str, Any], output_file: Path) -> BuildResult:
    company_name = str(report_data["company_name"])
    period_label = str(report_data["period_label"])
    receipts_rows = list(report_data["receipts_rows"])
    sales_rows = list(report_data["sales_rows"])

    output_file = make_unique_output_path(output_file.resolve())
    output_file.parent.mkdir(parents=True, exist_ok=True)
    build_template(output_file=output_file, company_name=company_name, period=period_label)

    wb = load_workbook(output_file)
    ws_receipts = wb.active
    ws_receipts.title = "Поступление"
    ws_sales = wb.copy_worksheet(ws_receipts)
    ws_sales.title = "Реализация"
    attach_header_image(ws_sales)

    for ws in (ws_receipts, ws_sales):
        ws.freeze_panes = "A6"
        safe_unmerge(ws, "A12:J12")
        safe_unmerge(ws, "A14:J14")
        safe_unmerge(ws, "A22:J22")
        drop_second_section(ws)

    write_section_to_template(
        ws_receipts,
        first_data_row=RECEIPTS_FIRST_DATA_ROW,
        total_row=RECEIPTS_TOTAL_ROW,
        rows=receipts_rows,
        total_label="ПОСТУПЛЕНИЕ",
    )
    write_section_to_template(
        ws_sales,
        first_data_row=RECEIPTS_FIRST_DATA_ROW,
        total_row=RECEIPTS_TOTAL_ROW,
        rows=sales_rows,
        total_label="РЕАЛИЗАЦИЯ",
    )

    ws_sales.cell(row=RECEIPTS_TITLE_ROW, column=1, value="Реализация")
    wb.save(output_file)
    return BuildResult(
        receipts_count=len(receipts_rows),
        sales_count=len(sales_rows),
        output_file=output_file,
        company_name=company_name,
    )


def build_local_report(receipts_file: Path, sales_file: Path, output_file: Path) -> BuildResult:
    payload = build_payload_from_files(receipts_file, sales_file)
    report_data = build_report_data_from_payload(payload)
    return render_report_data_to_excel(report_data, output_file)


def build_local_reports(receipts_file: Path, sales_file: Path, output_file: Path) -> MultiBuildResult:
    payload = build_payload_from_files(receipts_file, sales_file)
    report_data = build_report_data_from_payload(payload)
    grouped_reports = split_report_rows_by_comment_group(report_data)

    builds: list[BuildResult] = []
    for comment_group, grouped_report_data in grouped_reports:
        current_output = build_group_output_path(output_file, comment_group)
        builds.append(render_report_data_to_excel(grouped_report_data, current_output))

    return MultiBuildResult(
        company_name=str(report_data["company_name"]),
        builds=tuple(builds),
    )
