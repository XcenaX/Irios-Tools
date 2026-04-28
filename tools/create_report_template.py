from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import points_to_pixels

try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError:  # pragma: no cover
    Image = None
    ImageDraw = None
    ImageFont = None


if getattr(sys, "frozen", False):
    BASE_DIR = Path(getattr(sys, "_MEIPASS", Path(sys.executable).resolve().parent))
else:
    BASE_DIR = Path(__file__).resolve().parents[1]
OUTPUT_FILE = BASE_DIR / "report_template.xlsx"
GENERATED_DIR = BASE_DIR / "_generated"
HEADER_IMAGE = GENERATED_DIR / "header_banner.png"
BG_IMAGE = BASE_DIR / "assets" / "bg.png"
LOGO1_IMAGE = BASE_DIR / "assets" / "logo1.png"
LOGO2_IMAGE = BASE_DIR / "assets" / "logo2.png"
SHEET_NAME = "Шаблон"

NAVY = "1C1F52"
GOLD = "C9A55A"
PANEL = "F4F2FB"
PANEL_BORDER = "D7D0E7"
TEXT_DARK = "222222"
TEXT_LIGHT = "FFFFFF"
ACCENT_FILL = "F7EFD8"

THIN_PANEL = Side(style="thin", color=PANEL_BORDER)
THIN_DARK = Side(style="thin", color="4A4F78")

PANEL_BORDER_STYLE = Border(
    left=THIN_PANEL, right=THIN_PANEL, top=THIN_PANEL, bottom=THIN_PANEL
)
HEADER_BORDER_STYLE = Border(
    left=THIN_DARK, right=THIN_DARK, top=THIN_DARK, bottom=THIN_DARK
)


def apply_border_range(
    ws, start_row: int, end_row: int, start_col: int, end_col: int, border: Border
) -> None:
    for row in ws.iter_rows(
        min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col
    ):
        for cell in row:
            cell.border = border


def apply_fill_range(
    ws, start_row: int, end_row: int, start_col: int, end_col: int, fill: PatternFill
) -> None:
    for row in ws.iter_rows(
        min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col
    ):
        for cell in row:
            cell.fill = fill


def set_column_widths(ws) -> None:
    widths = {
        "A": 6,
        "B": 18,
        "C": 18,
        "D": 18,
        "E": 14,
        "F": 16,
        "G": 12,
        "H": 12,
        "I": 12,
        "J": 27.1429,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def column_width_to_pixels(width: float | None) -> int:
    if width is None:
        width = 8.43
    return int(width * 7 + 5)


def sheet_range_size_pixels(
    ws,
    *,
    start_col: int,
    end_col: int,
    start_row: int,
    end_row: int,
) -> tuple[int, int]:
    total_width = 0
    for column_index in range(start_col, end_col + 1):
        letter = get_column_letter(column_index)
        total_width += column_width_to_pixels(ws.column_dimensions[letter].width)

    total_height = 0
    for row_index in range(start_row, end_row + 1):
        height = ws.row_dimensions[row_index].height or 15
        total_height += points_to_pixels(height)

    return total_width, total_height


def fit_logo(image: Image.Image, max_width: int, max_height: int) -> Image.Image:
    ratio = min(max_width / image.width, max_height / image.height)
    new_size = (max(1, int(image.width * ratio)), max(1, int(image.height * ratio)))
    return image.resize(new_size, Image.LANCZOS)


def load_font(size: int, *, bold: bool = False) -> ImageFont.ImageFont:
    candidates = []
    if bold:
        candidates.extend(
            [
                r"C:\Windows\Fonts\arialbd.ttf",
                r"C:\Windows\Fonts\segoeuib.ttf",
                r"C:\Windows\Fonts\timesbd.ttf",
            ]
        )
    else:
        candidates.extend(
            [
                r"C:\Windows\Fonts\arial.ttf",
                r"C:\Windows\Fonts\segoeui.ttf",
                r"C:\Windows\Fonts\times.ttf",
            ]
        )

    for path in candidates:
        font_path = Path(path)
        if font_path.exists():
            return ImageFont.truetype(str(font_path), size)
    return ImageFont.load_default()


def ensure_header_asset(company_name: str = "", period: str = "") -> Path | None:
    if Image is None or ImageDraw is None or ImageFont is None:
        return None
    if not (BG_IMAGE.exists() and LOGO1_IMAGE.exists() and LOGO2_IMAGE.exists()):
        return None

    GENERATED_DIR.mkdir(exist_ok=True)

    banner_width = 2050
    banner_height = 210
    bottom_band_height = 62
    divider_y = banner_height - bottom_band_height

    with Image.open(BG_IMAGE).convert("RGBA") as bg:
        background = bg.resize((banner_width, banner_height), Image.LANCZOS)

    draw = ImageDraw.Draw(background)
    draw.rectangle([(0, divider_y - 1), (banner_width, divider_y + 10)], fill=(201, 165, 90, 255))

    with Image.open(LOGO1_IMAGE).convert("RGBA") as logo1:
        left_logo = fit_logo(logo1, 360, 96)
    with Image.open(LOGO2_IMAGE).convert("RGBA") as logo2:
        right_logo = fit_logo(logo2, 280, 80)

    background.alpha_composite(left_logo, (22, 26))
    background.alpha_composite(right_logo, (470, 33))

    title_font = load_font(40, bold=True)
    label_font = load_font(22)
    value_font = load_font(18, bold=True)

    upper_center_y = divider_y // 2

    draw.text(
        (1630, upper_center_y),
        "КОНТРОЛЬ ПЕРВИЧНЫХ ДОКУМЕНТОВ",
        font=title_font,
        fill=(255, 255, 255, 255),
        anchor="rm",
    )

    baseline_y = 180
    draw.text(
        (820, baseline_y),
        "Компания:",
        font=label_font,
        fill=(232, 232, 240, 255),
        anchor="rm",
    )
    draw.line((845, baseline_y + 4, 1105, baseline_y + 4), fill=(245, 245, 245, 255), width=2)
    if company_name:
        draw.text(
            (975, baseline_y - 10),
            company_name,
            font=value_font,
            fill=(255, 255, 255, 245),
            anchor="mm",
        )

    draw.text((1128, baseline_y - 3), "|", font=label_font, fill=(215, 215, 225, 255), anchor="mm")

    draw.text(
        (1265, baseline_y),
        "Период:",
        font=label_font,
        fill=(232, 232, 240, 255),
        anchor="rm",
    )
    draw.line((1290, baseline_y + 4, 1555, baseline_y + 4), fill=(245, 245, 245, 255), width=2)
    if period:
        draw.text(
            (1422, baseline_y - 10),
            period,
            font=value_font,
            fill=(255, 255, 255, 245),
            anchor="mm",
        )

    background.convert("RGB").save(HEADER_IMAGE)
    return HEADER_IMAGE


def add_banner(ws, company_name: str = "", period: str = "") -> None:
    banner_fill = PatternFill("solid", fgColor=NAVY)

    ws.merge_cells("A1:J3")
    apply_fill_range(ws, 1, 3, 1, 10, banner_fill)
    apply_border_range(ws, 1, 3, 1, 10, HEADER_BORDER_STYLE)

    asset_path = ensure_header_asset(company_name=company_name, period=period)
    if asset_path is not None:
        image = XLImage(str(asset_path))
        image.width, image.height = sheet_range_size_pixels(
            ws,
            start_col=1,
            end_col=10,
            start_row=1,
            end_row=3,
        )
        ws.add_image(image, "A1")


def add_section(
    ws,
    *,
    start_row: int,
    title: str,
    total_label: str,
    data_rows: int = 6,
) -> int:
    panel_fill = PatternFill("solid", fgColor=PANEL)
    header_fill = PatternFill("solid", fgColor=NAVY)
    accent_fill = PatternFill("solid", fgColor=ACCENT_FILL)

    section_end_row = start_row + data_rows + 3
    apply_fill_range(ws, start_row, section_end_row, 1, 10, panel_fill)
    apply_border_range(ws, start_row, section_end_row, 1, 10, PANEL_BORDER_STYLE)

    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(name="Calibri", size=21, color=TEXT_DARK)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    header_row = start_row + 1
    headers = {
        1: "№",
        2: "КОНТРАГЕНТ",
        5: "ДАТА",
        6: "СУММА",
        7: "ВИД ОПЕРАЦИИ",
        10: "УЧЁТНЫЙ № (1С)",
    }
    ws.merge_cells(start_row=header_row, start_column=2, end_row=header_row, end_column=4)
    ws.merge_cells(start_row=header_row, start_column=7, end_row=header_row, end_column=9)
    for column, value in headers.items():
        cell = ws.cell(row=header_row, column=column, value=value)
        cell.font = Font(name="Calibri", size=12, bold=True, color=TEXT_LIGHT)
        cell.fill = header_fill
        cell.border = HEADER_BORDER_STYLE
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for column in (3, 4, 8, 9):
        ws.cell(row=header_row, column=column).fill = header_fill
        ws.cell(row=header_row, column=column).border = HEADER_BORDER_STYLE

    first_data_row = header_row + 1
    for row_idx in range(first_data_row, first_data_row + data_rows):
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=4)
        ws.merge_cells(start_row=row_idx, start_column=7, end_row=row_idx, end_column=9)
        for column in range(1, 11):
            cell = ws.cell(row=row_idx, column=column)
            cell.border = PANEL_BORDER_STYLE
            cell.alignment = Alignment(
                horizontal="left" if column == 2 else "center",
                vertical="center",
            )
            if column == 6:
                cell.fill = accent_fill
                cell.number_format = '#,##0 "₸"'
        ws.cell(row=row_idx, column=1, value=row_idx - first_data_row + 1)
        ws.cell(row=row_idx, column=5, value="dd.mm.yyyy")

    total_row = first_data_row + data_rows
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=10)
    total_cell = ws.cell(
        row=total_row,
        column=1,
        value=f"ИТОГО: {total_label}: {{count}} документов / {{amount}} ₸",
    )
    total_cell.font = Font(name="Calibri", size=13, bold=True, color=TEXT_DARK)
    total_cell.alignment = Alignment(horizontal="left", vertical="center")

    return total_row


def build_template(
    output_file: Path,
    company_name: str = "",
    period: str = "",
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"

    set_column_widths(ws)

    for row_idx, height in {
        1: 36,
        2: 34,
        3: 15,
        4: 28,
        5: 26,
        14: 28,
        15: 26,
    }.items():
        ws.row_dimensions[row_idx].height = height

    add_banner(ws, company_name=company_name, period=period)
    first_total_row = add_section(
        ws,
        start_row=4,
        title="Поступление",
        total_label="ПОСТУПЛЕНИЕ",
    )
    add_section(
        ws,
        start_row=first_total_row + 2,
        title="Реализация",
        total_label="РЕАЛИЗАЦИЯ",
    )

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = "$1:$5"
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35

    wb.save(output_file.resolve())
    return output_file.resolve()


if __name__ == "__main__":
    print(build_template(OUTPUT_FILE))
